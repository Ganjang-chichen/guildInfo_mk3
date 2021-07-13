import openpyxl
from openpyxl.chart import BarChart3D, Reference, LineChart
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import PatternFill, Border, Font, Side
from openpyxl.formatting.rule import CellIsRule
import os
import math
import json
from datetime import date

#import week_calc
from function import week_calc

SHEET_VERSION = 1.3

file_path = "./ExcelFile/test.xlsx"

World = "29"
guild_name = "Lune"
guild_data = {}
guild_data_path = "./jsondata/" + World + "/" + guild_name + "/charData.json"
guild_imgData_path = "./img/" + World + "/" + guild_name + "/"

def get_chardata() :
    if os.path.exists(guild_data_path):
        global guild_data
        with open(guild_data_path, 'r', encoding="UTF8") as f:
            json_data = json.load(f)
        guild_data = json_data
        return True
    else :
        return False


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def createExcel_if_not_exist() :   
    if os.path.exists(file_path) :
        wb = openpyxl.load_workbook(file_path)
    else :
        wb = openpyxl.Workbook()
        wb.save(file_path)
    return wb

def setting_summary_sheet(wb) :
    sheet_names = wb.sheetnames
    data_count = len(sheet_names) - 1
    summary_sheet = wb["Sheet"]

    summary_sheet['O2'].value = "version"
    summary_sheet['O3'].value = SHEET_VERSION

    summary_sheet["C64"] = "날짜"
    summary_sheet["D64"] = "참가수"
    summary_sheet["E64"] = "주간미션 합계"
    summary_sheet["F64"] = "수로 합계"
    summary_sheet["G64"] = "플래그 합계"

    i = 0
    for sheet_name in sheet_names :
        if sheet_name == "Sheet" :
            continue
        else :
            sundayInt = int(sheet_name.split("-")[1])
            sunday = date((sundayInt // 10000), (((sundayInt - (sundayInt // 10000) * 10000)) // 100), (sundayInt % 100) )
            summary_sheet["F" + str(65 + i)] = sunday
            summary_sheet["G" + str(65 + i)] = "=AVERAGE('" + sheet_name + "'!H3, '" + sheet_name + "'!I3, '" + sheet_name + "'!J3)"
            summary_sheet["H" + str(65 + i)] = "='" + sheet_name + "'!H6"
            summary_sheet["I" + str(65 + i)] = "='" + sheet_name + "'!I6"
            summary_sheet["J" + str(65 + i)] = "='" + sheet_name + "'!J6"
            i = i + 1

    sunday_dateList = Reference(summary_sheet, min_col=3, min_row=65, max_row=(64 + data_count))

    average_entry_data = Reference(summary_sheet, min_col=4, max_col=4, min_row=64, max_row=(64 + data_count))
    average_entry_chart = LineChart()
    average_entry_chart.title = "평균 참가자수"
    average_entry_chart.style = 12
    average_entry_chart.y_axis.title = "참가자수"
    average_entry_chart.y_axis.crossAx = 500
    average_entry_chart.x_axis = DateAxis(crossAx=100)
    average_entry_chart.x_axis.number_format = 'd-mmm'
    average_entry_chart.x_axis.majorTimeUnit = "days"
    average_entry_chart.x_axis.title = "Date"
    average_entry_chart.add_data(average_entry_data, titles_from_data=True)
    average_entry_chart.set_categories(sunday_dateList)
    summary_sheet.add_chart(average_entry_chart, "F2")

    weekly_mission_data = Reference(summary_sheet, min_col=5, max_col=5, min_row=64, max_row=(64 + data_count))
    weekly_mission_chart = LineChart()
    weekly_mission_chart.title = "주간 미션 합계"
    weekly_mission_chart.style = 12
    weekly_mission_chart.y_axis.title = "점수"
    weekly_mission_chart.y_axis.crossAx = 500
    weekly_mission_chart.x_axis = DateAxis(crossAx=100)
    weekly_mission_chart.x_axis.number_format = 'd-mmm'
    weekly_mission_chart.x_axis.majorTimeUnit = "days"
    weekly_mission_chart.x_axis.title = "Date"
    weekly_mission_chart.add_data(weekly_mission_data, titles_from_data=True)
    weekly_mission_chart.set_categories(sunday_dateList)
    summary_sheet.add_chart(weekly_mission_chart, "F17")

    suro_sum_data = Reference(summary_sheet, min_col=6, max_col=6, min_row=64, max_row=(64 + data_count))
    suro_sum_chart = LineChart()
    suro_sum_chart.title = "수로 합계"
    suro_sum_chart.style = 12
    suro_sum_chart.y_axis.title = "점수"
    suro_sum_chart.y_axis.crossAx = 500
    suro_sum_chart.x_axis = DateAxis(crossAx=100)
    suro_sum_chart.x_axis.number_format = 'd-mmm'
    suro_sum_chart.x_axis.majorTimeUnit = "days"
    suro_sum_chart.x_axis.title = "Date"
    suro_sum_chart.add_data(suro_sum_data, titles_from_data=True)
    suro_sum_chart.set_categories(sunday_dateList)
    summary_sheet.add_chart(suro_sum_chart, "F32")

    flag_sum_data = Reference(summary_sheet, min_col=7, max_col=7, min_row=64, max_row=(64 + data_count))
    flag_sum_chart = LineChart()
    flag_sum_chart.title = "플래그 합계"
    flag_sum_chart.style = 12
    flag_sum_chart.y_axis.title = "점수"
    flag_sum_chart.y_axis.crossAx = 500
    flag_sum_chart.x_axis = DateAxis(crossAx=100)
    flag_sum_chart.x_axis.number_format = 'd-mmm'
    flag_sum_chart.x_axis.majorTimeUnit = "days"
    flag_sum_chart.x_axis.title = "Date"
    flag_sum_chart.add_data(flag_sum_data, titles_from_data=True)
    flag_sum_chart.set_categories(sunday_dateList)
    summary_sheet.add_chart(flag_sum_chart, "F47")

    wb.save(file_path)
    
    

def createSheet_if_not_exist(wb, sheet_name, guild_data) :
    wb.active
    if sheet_name in wb.sheetnames :
        return wb[sheet_name]
    else :
        sheet = 0
        if len(wb.sheetnames) == 1 :
            sheet = wb.create_sheet(sheet_name)
        else :
            sheet_insert_idx = 1
            while sheet_insert_idx < len(wb.sheetnames) :
                if int(wb.sheetnames[sheet_insert_idx].split('-')[0]) > int(sheet_name.split('-')[0]) :
                    break
                sheet_insert_idx += 1
            sheet = wb.create_sheet(sheet_name, sheet_insert_idx)    

        
        sheet.cell(row=1, column=1).value = "케릭터 명"
        sheet.cell(row=1, column=2).value = "주간 미션"
        sheet.cell(row=1, column=3).value = "수로"
        sheet.cell(row=1, column=4).value = "플래그"

        sheet.merge_cells(start_row=1, end_row=1, start_column=7, end_column=10)
        sheet.cell(row=1, column=7).value = "이번주"

        sheet.cell(row=2, column=7).value = "구분"
        sheet.cell(row=3, column=7).value = "참가자 수"
        sheet.cell(row=4, column=7).value = "점수 평균"
        sheet.cell(row=5, column=7).value = "참가자 평균"
        sheet.cell(row=6, column=7).value = "점수 합산"

        sheet.cell(row=2, column=8).value = "주간 미션"
        sheet.cell(row=2, column=9).value = "수로 "
        sheet.cell(row=2, column=10).value = "플래그"

        sheet.cell(row=3, column=8).value = '=COUNTIF(B2:B201,">0")'
        sheet.cell(row=3, column=9).value = '=COUNTIF(C2:C201,">0")'
        sheet.cell(row=3, column=10).value = '=COUNTIF(D2:D201,">0")'

        sheet.cell(row=4, column=8).value = '=AVERAGE(B2:B201)'
        sheet.cell(row=4, column=9).value = '=AVERAGE(C2:C201)'
        sheet.cell(row=4, column=10).value = '=AVERAGE(D2:D201)'

        sheet.cell(row=5, column=8).value = '=AVERAGEIF(B2:B201,">0")'
        sheet.cell(row=5, column=9).value = '=AVERAGEIF(C2:C201,">0")'
        sheet.cell(row=5, column=10).value = '=AVERAGEIF(D2:D201,">0")'

        sheet.cell(row=6, column=8).value = '=SUM(B2:B201)'
        sheet.cell(row=6, column=9).value = '=SUM(C2:C201)'
        sheet.cell(row=6, column=10).value = '=SUM(D2:D201)'
   

        if len(wb.sheetnames) != 1 :
            week_int = 0
            while True :
                temp = week_calc.get_weekinfo(week_int)
                temp_str = week_calc.get_week_range(temp)
                
                if wb.sheetnames[len(wb.sheetnames) - 1] == temp_str:
                    break
                week_int += 1

            last_sheet_name = week_calc.get_week_range(week_calc.get_weekinfo(week_int + 1))
            sheet.merge_cells(start_row=1, end_row=1, start_column=12, end_column=15)
            sheet.cell(row=1, column=12).value = "저번주"

            sheet.cell(row=2, column=12).value = "구분"
            sheet.cell(row=3, column=12).value = "참가자 수"
            sheet.cell(row=4, column=12).value = "점수 평균"
            sheet.cell(row=5, column=12).value = "참가자 평균"
            sheet.cell(row=6, column=12).value = "점수 합산"

            sheet.cell(row=2, column=13).value = "주간 미션"
            sheet.cell(row=2, column=14).value = "수로 "
            sheet.cell(row=2, column=15).value = "플래그"

            sheet.cell(row=3, column=13).value = "='" + last_sheet_name + "'!H3"
            sheet.cell(row=3, column=14).value = "='" + last_sheet_name + "'!I3"
            sheet.cell(row=3, column=15).value = "='" + last_sheet_name + "'!J3"

            sheet.cell(row=4, column=13).value = "='" + last_sheet_name + "'!H4"
            sheet.cell(row=4, column=14).value = "='" + last_sheet_name + "'!I4"
            sheet.cell(row=4, column=15).value = "='" + last_sheet_name + "'!J4"

            sheet.cell(row=5, column=13).value = "='" + last_sheet_name + "'!H5"
            sheet.cell(row=5, column=14).value = "='" + last_sheet_name + "'!I5"
            sheet.cell(row=5, column=15).value = "='" + last_sheet_name + "'!J5"

            sheet.cell(row=6, column=13).value = "='" + last_sheet_name + "'!H6"
            sheet.cell(row=6, column=14).value = "='" + last_sheet_name + "'!I6"
            sheet.cell(row=6, column=15).value = "='" + last_sheet_name + "'!J6"

            sheet.cell(row=9, column=7).value = "주간 미션"
            sheet.cell(row=10, column=7).value = "참가자수"
            sheet.cell(row=11, column=7).value = "점수 평균"
            sheet.cell(row=12, column=7).value = "참가자 평균"
            sheet.cell(row=13, column=7).value = "점수 합산"

            sheet.cell(row=9, column=8).value = "이번주"
            sheet.cell(row=10, column=8).value = "=H3"
            sheet.cell(row=11, column=8).value = "=H4"
            sheet.cell(row=12, column=8).value = "=H5"
            sheet.cell(row=13, column=8).value = "=H6"

            sheet.cell(row=9, column=9).value = "저번주"
            sheet.cell(row=10, column=9).value = "=M3"
            sheet.cell(row=11, column=9).value = "=M4"
            sheet.cell(row=12, column=9).value = "=M5"
            sheet.cell(row=13, column=9).value = "=M6"

            weekly_mission_data = Reference(sheet, min_col=8, max_col=9, min_row=9, max_row=13)
            weekly_mission_title = Reference(sheet, min_col=7, min_row=10, max_row=13)
            weekly_mission_chart = BarChart3D()
            weekly_mission_chart.title = "주간 미션 비교"
            weekly_mission_chart.add_data(data=weekly_mission_data, titles_from_data=True)
            weekly_mission_chart.set_categories(weekly_mission_title)
            sheet.add_chart(weekly_mission_chart, "G17")

            sheet.cell(row=9, column=11).value = "수로"
            sheet.cell(row=10, column=11).value = "참가자수"
            sheet.cell(row=11, column=11).value = "점수 평균"
            sheet.cell(row=12, column=11).value = "참가자 평균"
            sheet.cell(row=13, column=11).value = "점수 합산"

            sheet.cell(row=9, column=12).value = "이번주"
            sheet.cell(row=10, column=12).value = "=I3"
            sheet.cell(row=11, column=12).value = "=I4"
            sheet.cell(row=12, column=12).value = "=I5"
            sheet.cell(row=13, column=12).value = "=I6"

            sheet.cell(row=9, column=13).value = "저번주"
            sheet.cell(row=10, column=13).value = "=N3"
            sheet.cell(row=11, column=13).value = "=N4"
            sheet.cell(row=12, column=13).value = "=N5"
            sheet.cell(row=13, column=13).value = "=N6"

            suro_data = Reference(sheet, min_col=12, max_col=13, min_row=9, max_row=13)
            suro_title = Reference(sheet, min_col=11, min_row=10, max_row=13)
            suro_chart = BarChart3D()
            suro_chart.title = "수로 비교"
            suro_chart.add_data(data=suro_data, titles_from_data=True)
            suro_chart.set_categories(suro_title)
            sheet.add_chart(suro_chart, "G32")

            sheet.cell(row=9, column=15).value = "플래그"
            sheet.cell(row=10, column=15).value = "참가자수"
            sheet.cell(row=11, column=15).value = "점수 평균"
            sheet.cell(row=12, column=15).value = "참가자 평균"
            sheet.cell(row=13, column=15).value = "점수 합산"

            sheet.cell(row=9, column=16).value = "이번주"
            sheet.cell(row=10, column=16).value = "=J3"
            sheet.cell(row=11, column=16).value = "=J4"
            sheet.cell(row=12, column=16).value = "=J5"
            sheet.cell(row=13, column=16).value = "=J6"

            sheet.cell(row=9, column=17).value = "저번주"
            sheet.cell(row=10, column=17).value = "=O3"
            sheet.cell(row=11, column=17).value = "=O4"
            sheet.cell(row=12, column=17).value = "=O5"
            sheet.cell(row=13, column=17).value = "=O6"

            flag_data = Reference(sheet, min_col=16, max_col=17, min_row=9, max_row=13)
            flag_title = Reference(sheet, min_col=15, min_row=10, max_row=13)
            flag_chart = BarChart3D()
            flag_chart.title = "플래그 비교"
            flag_chart.add_data(data=flag_data, titles_from_data=True)
            flag_chart.set_categories(flag_title)
            sheet.add_chart(flag_chart, "G47")

        i = 0
        for rowData in guild_data['charData'] :
            sheet.cell(row=(i + 2), column=1).value = rowData['name']
            sheet.cell(row=(i + 2), column=2).value = 0
            sheet.cell(row=(i + 2), column=3).value = 0
            sheet.cell(row=(i + 2), column=4).value = 0
            i = i + 1

            rate_fill = PatternFill(start_color='FF5348', end_color='FF5348', fill_type='solid')
            rate_font = Font(color="FFFFFF")
            rate_thin = Side(border_style='thin', color="000000")
            rate_border = Border(top=rate_thin, left=rate_thin, right=rate_thin, bottom=rate_thin)
            sheet.conditional_formatting.add('B' + str(i + 1) + ":D" + str(i + 1), CellIsRule(operator='lessThan', formula=[1], fill=rate_fill, font=rate_font, border=rate_border))

        setting_summary_sheet(wb)

        wb.save(file_path)
        return sheet

if __name__ == "__main__" :
    createFolder("./testExcelFile")
    get_chardata()

    wb = createExcel_if_not_exist()
    weekinfo = week_calc.get_weekinfo(0)
    sheet_name = (  weekinfo["first"]["year"] + weekinfo["first"]["month"] + weekinfo["first"]["day"] + "-" + 
                    weekinfo["last"]["year"] + weekinfo["last"]["month"] + weekinfo["last"]["day"])
    print(sheet_name)
    createSheet_if_not_exist(wb, sheet_name, guild_data)
    print(wb.sheetnames)
    sheet_names_reverse = wb.sheetnames[::-1]
    sheet_names_reverse.pop()
    print(wb.sheetnames)
    print(sheet_names_reverse)
